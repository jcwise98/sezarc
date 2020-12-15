from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from PIL import ImageTk, Image
import grapher
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


# Useful Vars
global is_imported
global first_column
is_imported = False  # keep track of whether or not a spreadsheet has been imported


# Function for opening the
# file explorer window
def get_spreadsheet():
    global is_imported
    global first_column
    filename = filedialog.askopenfilename(initialdir="",
                                          title="Select a File",
                                          filetypes=(("Excel files",
                                                      "*.xlsx*"),
                                                     ("all files",
                                                      "*.*")))

    wb = Workbook()
    # wb = load_workbook('MarinelandThru30June2020_FormattedWithDepth.xlsx')
    wb = load_workbook(filename)
    ws = wb.active

    first_column = ws['A']
    spreadsheet_info.config(text="Imported Spreadsheet: " + filename)
    is_imported = True

    fig = grapher.heat_map_plot(filename)
    canvas_heat = FigureCanvas(fig)


# Function to display data from spreadsheet
def get_list():
    if is_imported is True:
        list = ''
        for cell in first_column:
            list = f'{list + str(cell.value)}\n'

            coord_list.config(text=list)
    else:
        messagebox.showerror("Error", "No Spreadsheet Imported")


# Function to hide data from spreadsheet
def remove_list():
    coord_list.config(text="")


# Function to get image for enclosure
def get_image():
    imagename = filedialog.askopenfilename(initialdir="",
                                          title="Select a File",
                                          filetypes=(("Image files",
                                                      "*.png*"),
                                                     ("all files",
                                                      "*.*")))
    img = ImageTk.PhotoImage(Image.open(imagename))
    canvas.create_image(20, 20, anchor=NW, image=img)
    enclosure_image.image = img


def print_dev():
    messagebox.showinfo("Developers", "Farah Aljishi\nDerek Baum\nRyan Bonacquisti\nDebra Lymon\nNtsee Ndingwan\nJake Wise")
    # print("Farah Aljishi\nDerek Baum\nRyan Bonacquisti\nDebra Lymon\nNtsee Ndingwan\nJake Wise")


root = Tk()
root.title('ZooMonitor Data Mapper')
root.geometry("1440x810")

menu = Menu(root)
root.config(menu=menu)

# Define Menus
fileMenu = Menu(menu)
fileMenu = Menu(menu, tearoff=0)
editMenu = Menu(menu)
editMenu = Menu(menu, tearoff=0)
viewMenu = Menu(menu)
viewMenu = Menu(menu, tearoff=0)
aboutMenu = Menu(menu)
aboutMenu = Menu(menu, tearoff=0)
menu.add_cascade(label="File", menu=fileMenu)
menu.add_cascade(label="Edit", menu=editMenu)
menu.add_cascade(label="View", menu=viewMenu)
menu.add_cascade(label="About", menu=aboutMenu)

# File Menu Options
fileMenu.add_command(label="Import Spreadsheet", command=get_spreadsheet)
fileMenu.add_command(label="Import Habitat", command=get_image)

# Edit Menu Options

# View Menu Options
viewMenu.add_command(label="Display Coordinates", command=get_list)
viewMenu.add_command(label="Hide Coordinates", command=remove_list)

# About Menu Options
aboutMenu.add_command(label="Developers", command=print_dev)

spreadsheet_info = Label(root, text="")
spreadsheet_info.pack()

coord_list = Label(root, text="")
coord_list.pack()

canvas = Canvas(root, width=600, height=600)
canvas.pack()
enclosure_image = Label(image="")
enclosure_image.pack()

root.mainloop()
