import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import PIL.Image
import PIL.ImageTk
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


LARGE_FONT= ("Verdana", 12)



class ZooMapper(tk.Tk):


    def __init__(self, *args, **kwargs):

        tk.Tk.__init__(self, *args, **kwargs)

        tk.Tk.iconbitmap(self, default="clienticon.ico")
        tk.Tk.wm_title(self, "ZooMonitor Data Mapper")
        tk.Tk.wm_geometry(self, "1440x810")

        menu = Menu(self)
        tk.Tk.config(self, menu=menu)

        # Define Menus
        fileMenu = Menu(menu)
        fileMenu = Menu(menu, tearoff=0)
        editMenu = Menu(menu)
        editMenu = Menu(menu, tearoff=0)
        viewMenu = Menu(menu)
        viewMenu = Menu(menu, tearoff=0)
        aboutMenu = Menu(menu)
        aboutMenu = Menu(menu, tearoff=0)
        menu.add_cascade(label="File", menu=fileMenu)
        menu.add_cascade(label="Edit", menu=editMenu)
        menu.add_cascade(label="View", menu=viewMenu)
        menu.add_cascade(label="About", menu=aboutMenu)

        # File Menu Options
        fileMenu.add_command(label="Import Spreadsheet", command=self.get_spreadsheet)
        fileMenu.add_command(label="Import Habitat", command=self.get_image)

        # Edit Menu Options

        # View Menu Options
        viewMenu.add_command(label="Display Coordinates", command=self.get_list)
        viewMenu.add_command(label="Hide Coordinates", command=self.remove_list)

        # About Menu Options
        aboutMenu.add_command(label="Developers", command=self.print_dev)


        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (StartPage, PageOne, PageTwo, PageThree):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)

        # Function to get image for enclosure
    def get_image(self):
        imagename = filedialog.askopenfilename(initialdir="",
                                               title="Select a File",
                                               filetypes=(("Image files",
                                                           "*.png*"),
                                                          ("all files",
                                                           "*.*")))
        img = PIL.ImageTk.PhotoImage(PIL.Image.open(imagename))
        canvas.create_image(0, 0, anchor=NW, image=img)
        enclosure_image.image = img

    # Function for opening the
    # file explorer window
    def get_spreadsheet(self):
        global is_imported
        global first_column
        filename = filedialog.askopenfilename(initialdir="",
                                              title="Select a File",
                                              filetypes=(("Excel files",
                                                          "*.xlsx*"),
                                                         ("all files",
                                                          "*.*")))

        wb = Workbook()
        # wb = load_workbook('MarinelandThru30June2020_FormattedWithDepth.xlsx')
        wb = load_workbook(filename)
        ws = wb.active

        first_column = ws['A']
        spreadsheet_info.config(text="Imported Spreadsheet: " + filename)
        is_imported = True

        fig = grapher.heat_map_plot(filename)
        canvas_heat = FigureCanvas(fig)


    # Function to display data from spreadsheet
    def get_list(self):
        if is_imported is True:
            list = ''
            for cell in first_column:
                list = f'{list + str(cell.value)}\n'

                coord_list.config(text=list)
        else:
            messagebox.showerror("Error", "No Spreadsheet Imported")


    # Function to hide data from spreadsheet
    def remove_list(self):
        coord_list.config(text="")


    def print_dev(self):
        messagebox.showinfo("Developers", "Farah Aljishi\nDerek Baum\nRyan Bonacquisti\nDebra Lymon\nNtsee Ndingwan\nJake Wise")
        # print("Farah Aljishi\nDerek Baum\nRyan Bonacquisti\nDebra Lymon\nNtsee Ndingwan\nJake Wise")

    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()

class StartPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        label = tk.Label(self, text="Start Page", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        button = ttk.Button(self, text="Visit Page 1",
                            command=lambda: controller.show_frame(PageOne))
        button.pack()

        button2 = ttk.Button(self, text="Visit Page 2",
                             command=lambda: controller.show_frame(PageTwo))
        button2.pack()

        button3 = ttk.Button(self, text="Graph Page",
                             command=lambda: controller.show_frame(PageThree))
        button3.pack()

        canvas = Canvas(self, width=600, height=600)
        canvas.pack()
        enclosure_image = Label(image="")
        enclosure_image.pack()


class PageOne(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text="Page One!!!", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        button1 = ttk.Button(self, text="Back to Home",
                             command=lambda: controller.show_frame(StartPage))
        button1.pack()

        button2 = ttk.Button(self, text="Page Two",
                             command=lambda: controller.show_frame(PageTwo))
        button2.pack()


class PageTwo(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text="Page Two!!!", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        button1 = ttk.Button(self, text="Back to Home",
                             command=lambda: controller.show_frame(StartPage))
        button1.pack()

        button2 = ttk.Button(self, text="Page One",
                             command=lambda: controller.show_frame(PageOne))
        button2.pack()


class PageThree(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text="Graph Page!", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        button1 = ttk.Button(self, text="Back to Home",
                             command=lambda: controller.show_frame(StartPage))
        button1.pack()

        f = Figure(figsize=(5,5), dpi=100)
        a = f.add_subplot(111)
        a.plot([1,2,3,4,5,6,7,8],[5,6,1,3,8,9,3,5])



        canvas = FigureCanvasTkAgg(f, self)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)

        toolbar = NavigationToolbar2Tk(canvas, self)
        toolbar.update()
        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)



app = ZooMapper()
app.mainloop()


